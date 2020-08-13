import openpyxl 
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill , colors, Font, Fill, Color, Border
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook

path = sys.argv[1]
wb = load_workbook(path)
i = len(wb.sheetnames)

RED = 'ff4d4d'
GREEN = '66ff66'
GRAY = 'bfbfbf'
BLACK = '000001'
WHITE = 'ffffff'

for sheet in wb.worksheets:
    ws = wb[wb.sheetnames[len(wb.sheetnames)-i]]
    
    for rows in ws.iter_rows():
        for cell in rows:
            if(str(cell.value).lower() == 'failed'):
                cell.fill=PatternFill(start_color= RED, end_color= RED, fill_type="solid")
                cell.font = Font(bold = True, color = GREEN, size = 20)
            elif(str(cell.value).lower() == 'passed'):
                cell.fill=PatternFill(start_color= GREEN, end_color= GREEN, fill_type="solid")
                cell.font = Font(bold = True, color = RED, size = 20)
            elif(str(cell.value).lower() == 'not performed'):
                cell.fill=PatternFill(start_color= GRAY, end_color= GRAY, fill_type="solid")
                cell.font = Font(bold = True, color = WHITE, size = 20)
    i += 1
    wb.save(path)